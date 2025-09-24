# Script: prioridades.py
import openpyxl
from openpyxl.styles import Alignment

def definir_prioridades():
    prioridades = []

    print("=== Definición de Prioridades ===")
    print("Ingresa tus prioridades y su nivel de importancia (1 = más alta, 5 = más baja).")
    print("Cuando termines, escribe 'fin' para dejar de ingresar prioridades.\n")

    while True:
        tarea = input("📌 Ingresa una prioridad (o escribe 'fin' para terminar): ")
        if tarea.lower() == "fin":
            break
        try:
            nivel = int(input(f"Nivel de importancia para '{tarea}' (1-5): "))
            if nivel < 1 or nivel > 5:
                print("⚠️ El nivel debe estar entre 1 y 5.")
                continue
            prioridades.append((nivel, tarea))
        except ValueError:
            print("⚠️ Debes ingresar un número entre 1 y 5.")

    # Ordenamos por nivel (más baja prioridad = número más alto)
    prioridades.sort(key=lambda x: x[0])

    # Guardamos en archivo TXT
    with open("prioridades.txt", "w", encoding="utf-8") as f:
        f.write("=== Lista de Prioridades ===\n\n")
        for i, (nivel, tarea) in enumerate(prioridades, start=1):
            f.write(f"{i}. {tarea} (Importancia: {nivel})\n")

    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prioridades"

    # Encabezados
    ws["A1"] = "Completado"
    ws["B1"] = "Prioridad"
    ws["C1"] = "Nivel de Importancia"

    # Ajustar alineación
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["B1"].alignment = Alignment(horizontal="center")
    ws["C1"].alignment = Alignment(horizontal="center")

    # Llenar datos con un checklist vacío en la columna A
    for i, (nivel, tarea) in enumerate(prioridades, start=2):
        ws[f"A{i}"] = "☐"   # Casilla sin marcar
        ws[f"B{i}"] = tarea
        ws[f"C{i}"] = nivel

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20

    wb.save("prioridades.xlsx")

    print("\n✅ Tus prioridades fueron guardadas en 'prioridades.txt' y 'prioridades.xlsx'.")

if __name__ == "__main__":
    definir_prioridades()

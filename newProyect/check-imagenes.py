import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from PIL import Image as PILImage


# ==============================
# CONFIGURACION
# ==============================

EXCEL_FILE = "table-excel.xlsx"
IMAGES_FOLDER = "imagenes"
OUTPUT_FILE = "table-excel-con-imagenes.xlsx"

COLUMNA_ITEM = 1
COLUMNA_DIBUJO = 2
COLUMNA_IMAGEN = 3

MAX_WIDTH = 140
MAX_HEIGHT = 90


# ==============================
# ABRIR EXCEL
# ==============================

wb = load_workbook(EXCEL_FILE)
ws = wb.active

# fijar ancho columna imagen
ws.column_dimensions['C'].width = 22


# ==============================
# INDEXAR IMAGENES
# ==============================

imagenes = {}

for file in os.listdir(IMAGES_FOLDER):

    if file.lower().endswith(".png"):

        match = re.match(r"(\d+)-?(B\.\d+)", file)

        if match:
            item = int(match.group(1))
            bnumero = match.group(2)

            clave = f"{item}-{bnumero}"
            imagenes[clave] = os.path.join(IMAGES_FOLDER, file)


# ==============================
# INSERTAR IMAGENES
# ==============================

for row in range(2, ws.max_row + 1):

    item = ws.cell(row=row, column=COLUMNA_ITEM).value
    dibujo = ws.cell(row=row, column=COLUMNA_DIBUJO).value

    if not item or not dibujo:
        continue

    clave = f"{int(item)}-{dibujo}"

    if clave in imagenes:

        ruta = imagenes[clave]

        # obtener tamaño real de imagen
        pil_img = PILImage.open(ruta)
        width, height = pil_img.size

        ratio = min(MAX_WIDTH / width, MAX_HEIGHT / height)

        new_width = int(width * ratio)
        new_height = int(height * ratio)

        img = Image(ruta)
        img.width = new_width
        img.height = new_height

        # posicion celda
        col = COLUMNA_IMAGEN - 1
        fila = row - 1

        # anclaje dentro de la celda
        anchor = TwoCellAnchor(
            _from=AnchorMarker(col=col, colOff=0, row=fila, rowOff=0),
            to=AnchorMarker(col=col+1, colOff=0, row=fila+1, rowOff=0),
        )

        img.anchor = anchor

        ws.add_image(img)

        # ajustar altura fila
        ws.row_dimensions[row].height = 70


# ==============================
# GUARDAR
# ==============================

wb.save(OUTPUT_FILE)

print("Proceso terminado. Excel generado:", OUTPUT_FILE)
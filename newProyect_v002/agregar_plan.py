import os
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook

# ============================================================
# CONFIGURACIÓN AUTOMÁTICA (usa la carpeta donde está el script)
# ============================================================
carpeta_base = os.path.dirname(os.path.abspath(__file__))

pdf_principal = os.path.join(carpeta_base, "PDF_unido.pdf")
archivo_excel = os.path.join(carpeta_base, "relacion_planos.xlsx")
carpeta_planes = os.path.join(carpeta_base, "plan_maquinado")
pdf_salida = os.path.join(carpeta_base, "PDF_unido_con_planes.pdf")

# ============================================================
# FUNCIÓN PARA BUSCAR PDF EN LA CARPETA
# Busca primero exacto: NOMBRE.pdf
# Si no lo encuentra, busca coincidencia ignorando mayúsculas/minúsculas
# ============================================================
def buscar_pdf(nombre_base, carpeta):
    nombre_esperado = f"{nombre_base}.pdf"
    ruta_exacta = os.path.join(carpeta, nombre_esperado)

    if os.path.exists(ruta_exacta):
        return ruta_exacta

    # Buscar ignorando mayúsculas/minúsculas
    for archivo in os.listdir(carpeta):
        if archivo.lower() == nombre_esperado.lower():
            return os.path.join(carpeta, archivo)

    return None

# ============================================================
# VALIDACIONES
# ============================================================
if not os.path.exists(pdf_principal):
    raise FileNotFoundError(f"No se encontró el PDF principal:\n{pdf_principal}")

if not os.path.exists(archivo_excel):
    raise FileNotFoundError(f"No se encontró el archivo Excel:\n{archivo_excel}")

if not os.path.exists(carpeta_planes):
    raise FileNotFoundError(f"No se encontró la carpeta de planes:\n{carpeta_planes}")

# ============================================================
# LEER PDF PRINCIPAL
# ============================================================
reader_principal = PdfReader(pdf_principal)
total_paginas = len(reader_principal.pages)

# ============================================================
# LEER EXCEL
# Espera:
# Columna A = número de página (Pag)
# Columna B = nombre del PDF (sin o con .pdf)
# ============================================================
wb = load_workbook(archivo_excel, data_only=True)
ws = wb.active

# Crear diccionario: pagina -> nombre_pdf
# Ejemplo: 1 -> "CONJUNTO"
mapa_paginas = {}

for fila in ws.iter_rows(min_row=2, values_only=True):  # empieza desde fila 2 (saltando encabezado)
    pag = fila[0]
    nombre_pdf = fila[1]

    if pag is None or nombre_pdf is None:
        continue

    try:
        pag = int(pag)
    except:
        print(f"⚠️ Página inválida en Excel: {pag}")
        continue

    nombre_pdf = str(nombre_pdf).strip()

    # Quitar .pdf si viene incluido
    if nombre_pdf.lower().endswith(".pdf"):
        nombre_pdf = nombre_pdf[:-4]

    mapa_paginas[pag] = nombre_pdf

# ============================================================
# CREAR PDF FINAL
# ============================================================
writer = PdfWriter()

print(f"\nProcesando PDF principal: {os.path.basename(pdf_principal)}")
print(f"Total de páginas: {total_paginas}\n")

for i in range(total_paginas):
    numero_pagina = i + 1

    # 1) Agregar página del PDF principal
    writer.add_page(reader_principal.pages[i])
    print(f"✔ Agregada página principal {numero_pagina}")

    # 2) Si existe un PDF asociado, insertarlo después
    if numero_pagina in mapa_paginas:
        nombre_pdf = mapa_paginas[numero_pagina]
        ruta_pdf_insertar = buscar_pdf(nombre_pdf, carpeta_planes)

        if ruta_pdf_insertar and os.path.exists(ruta_pdf_insertar):
            try:
                reader_insertar = PdfReader(ruta_pdf_insertar)
                paginas_insertadas = len(reader_insertar.pages)

                for pagina in reader_insertar.pages:
                    writer.add_page(pagina)

                print(f"   ➕ Insertado: {os.path.basename(ruta_pdf_insertar)} ({paginas_insertadas} página(s))")

            except Exception as e:
                print(f"   ❌ Error al leer '{nombre_pdf}': {e}")
        else:
            print(f"   ⚠ No se encontró el PDF para la página {numero_pagina}: {nombre_pdf}.pdf")

# ============================================================
# GUARDAR RESULTADO
# ============================================================
with open(pdf_salida, "wb") as f:
    writer.write(f)

print(f"\n✅ PDF final generado correctamente:")
print(pdf_salida)